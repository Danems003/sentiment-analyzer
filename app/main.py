from fastapi import FastAPI, Depends, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func
from datetime import datetime
from typing import Optional, Dict
import csv
from io import StringIO, BytesIO
from openpyxl import load_workbook
from dateutil import parser

from schemas import ReviewRequest, SentimentResponse
from sentiment_model import predict_sentiment
from kafka_producer import send_to_kafka
from database import SessionLocal, engine
from models import Review, Base

Base.metadata.create_all(bind=engine)

app = FastAPI()

# Разрешаем CORS (для фронтенда)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Получение сессии БД
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Анализ одного отзыва
@app.post("/analyze", response_model=SentimentResponse)
def analyze_sentiment(data: ReviewRequest, db: Session = Depends(get_db)):
    send_to_kafka("sentiment_topic", {"text": data.text})
    sentiment = predict_sentiment(data.text)

    review = Review(text=data.text, sentiment=sentiment)
    db.add(review)
    db.commit()
    db.refresh(review)

    return {"sentiment": sentiment}

# Загрузка CSV или XLSX файла
@app.post("/upload-file")
async def upload_file(file: UploadFile = File(...), db: Session = Depends(get_db)) -> Dict[str, int]:
    filename = file.filename.lower()
    stats = {"Положительный": 0, "Негативный": 0, "Нейтральный": 0}

    if filename.endswith(".csv"):
        contents = await file.read()
        decoded = contents.decode("utf-8")
        reader = csv.DictReader(StringIO(decoded))
        rows = [{"text": row.get("text", "").strip(), "created_at": row.get("created_at", "").strip()} for row in reader]

    elif filename.endswith(".xlsx"):
        contents = await file.read()
        workbook = load_workbook(filename=BytesIO(contents), data_only=True)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]
        try:
            text_idx = headers.index("text")
            date_idx = headers.index("created_at")
        except ValueError:
            return {"error": "Файл Excel должен содержать колонки 'text' и 'created_at'"}

        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append({
                "text": str(row[text_idx]).strip() if row[text_idx] else "",
                "created_at": str(row[date_idx]).strip() if row[date_idx] else ""
            })

    else:
        return {"error": "Файл должен быть .csv или .xlsx"}

    for row in rows:
        text = row["text"]
        created_at_str = row["created_at"]

        if not text or not created_at_str:
            continue

        try:
            created_at = parser.parse(created_at_str)
        except Exception as e:
            print(f"Ошибка парсинга даты: {created_at_str}, {e}")
            continue

        sentiment = predict_sentiment(text)
        review = Review(text=text, sentiment=sentiment, created_at=created_at)
        db.add(review)
        stats[sentiment] += 1

    db.commit()
    return stats

# Дашборд со статистикой по фильтрам
@app.get("/dashboard")
def get_dashboard(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    query = db.query(
        func.date(Review.created_at).label("date"),
        Review.sentiment,
        func.count().label("count")
    )

    if start_date:
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(Review.created_at >= start)
        except ValueError:
            return {"error": "Invalid start_date format. Use YYYY-MM-DD."}

    if end_date:
        try:
            end = datetime.strptime(end_date, "%Y-%m-%d")
            query = query.filter(Review.created_at <= end)
        except ValueError:
            return {"error": "Invalid end_date format. Use YYYY-MM-DD."}

    query = query.group_by(func.date(Review.created_at), Review.sentiment)
    query = query.order_by(func.date(Review.created_at))

    raw_results = query.all()

    # Собираем данные в структуру, подходящую для LineChart
    data_by_date = {}
    for date, sentiment, count in raw_results:
        date_str = date.strftime("%Y-%m-%d")
        if date_str not in data_by_date:
            data_by_date[date_str] = {"date": date_str, "Положительный": 0, "Негативный": 0, "Нейтральный": 0}
        data_by_date[date_str][sentiment] = count

    # Возвращаем отсортированный список словарей
    return list(data_by_date.values())




